# -*- coding: utf-8 -*-
"""
FastAPI · Exportador Excel Flora (DwC-SMA · 3 hojas)
----------------------------------------------------
• Endpoint: /export?campana_id=...
• Devuelve: {"download_url": ".../download/<archivo.xlsx>"}

Requisitos:
- FIREBASE_KEY_B64 (env var) → JSON service account en base64 (una línea)
- Plantilla XLSX en el repo: FormatoBiodiversidadMonitoreoYLineaBase_v5.2.xlsx
- Render Free: archivos en /tmp/downloads
"""

import os, re, json, base64, uuid, warnings
from pathlib import Path
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

from fastapi import FastAPI, Query, Request, HTTPException
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from openpyxl import load_workbook

import firebase_admin
from firebase_admin import credentials, firestore

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ──────────────────────────────── Paths & Const
ROOT_DIR      = Path(__file__).parent
TEMPLATE_PATH = ROOT_DIR / "FormatoBiodiversidadMonitoreoYLineaBase_v5.2.xlsx"
DOWNLOAD_DIR  = Path("/tmp/downloads")
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
LOCAL_TZ      = ZoneInfo("America/Santiago")

# ──────────────────────────────── Firebase Init (env)
B64 = os.environ.get("FIREBASE_KEY_B64")
if not B64:
    raise RuntimeError("FIREBASE_KEY_B64 env var is required")

cred_info = json.loads(base64.b64decode(B64))
if not firebase_admin._apps:
    firebase_admin.initialize_app(credentials.Certificate(cred_info))
db = firestore.client()

# ──────────────────────────────── FastAPI + CORS
app = FastAPI(title="Exporter Flora · DwC-SMA")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],            # en prod: lista exacta si usas credenciales
    allow_methods=["*"],            # incluye OPTIONS (preflight)
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

# Opcional: mount estático (no imprescindible si usas /download)
app.mount("/downloads", StaticFiles(directory=str(DOWNLOAD_DIR)), name="downloads")

# ──────────────────────────────── Utils
def _safe_filename(s: str) -> str:
    return re.sub(r"[^\w\-]+", "-", str(s)).strip("-") or "file"

def strip_tz(df: pd.DataFrame, tz=LOCAL_TZ) -> pd.DataFrame:
    if df.empty:
        return df
    for col in df.columns:
        if pd.api.types.is_datetime64tz_dtype(df[col]):
            df[col] = df[col].dt.tz_convert(tz).dt.tz_localize(None)
        elif df[col].dtype == "object":
            df[col] = df[col].map(
                lambda v: v.astimezone(tz).replace(tzinfo=None)
                if isinstance(v, datetime) and v.tzinfo else v
            )
    return df

def fetch_by_campana(collection: str, campana_id: str) -> pd.DataFrame:
    campana_id = str(campana_id).strip('"')
    ref = db.collection(collection).where("campanaID", "==", campana_id)
    rows = []
    for d in ref.stream():
        data = d.to_dict() or {}
        data["id"] = d.id
        rows.append(data)
    return strip_tz(pd.DataFrame(rows))

def get_lat(p):
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return None
    return p.get("latitude") if isinstance(p, dict) else getattr(p, "latitude", None)

def get_lon(p):
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return None
    return p.get("longitude") if isinstance(p, dict) else getattr(p, "longitude", None)

# ──────────────────────────────── Generación de Excel
def generar_excel_fauna_like(campana_id: str) -> Path:
    # 1) Leer datos de Firestore (flora: campana/estacion/registro)
    df_campana  = fetch_by_campana("campana",  campana_id)
    df_estacion = fetch_by_campana("estacion", campana_id)
    df_registro = fetch_by_campana("registro", campana_id)

    if df_campana.empty:
        raise HTTPException(status_code=404, detail="No hay documentos en 'campana' para ese campanaID.")
    # Nota: si estación o registro están vacíos, igual generamos el archivo con las hojas vacías.

    # ───── Hoja "Campaña" (DF plantilla)
    CAMPANA_MAP = {
        "Nombre campaña":    "name",
        "Número de campaña": "ncampana",
        "Fecha inicio":      "startDateCamp",
        "Fecha término":     "endDateCamp",
    }
    camp = df_campana.iloc[0].to_dict()
    _to_dt = lambda v: pd.to_datetime(v, errors="coerce")
    dt_ini = _to_dt(camp.get(CAMPANA_MAP["Fecha inicio"]))
    dt_fin = _to_dt(camp.get(CAMPANA_MAP["Fecha término"]))

    def _ymd(dt: Any):
        if pd.isna(dt):
            return None, None, None
        try:
            return int(dt.year), int(dt.month), int(dt.day)
        except Exception:
            return None, None, None

    anio_ini, mes_ini, dia_ini = _ymd(dt_ini)
    anio_fin, mes_fin, dia_fin = _ymd(dt_fin)

    df_campana_plantilla = pd.DataFrame([{
        "ID Campaña":               1,
        "Nombre campaña":           camp.get(CAMPANA_MAP["Nombre campaña"]),
        "Número de campaña":        camp.get(CAMPANA_MAP["Número de campaña"]),
        "Año inicio":               anio_ini,
        "Mes inicio":               mes_ini,
        "Día inicio":               dia_ini,
        "Año término":              anio_fin,
        "Mes término":              mes_fin,
        "Día término":              dia_fin,
        "Objetivo de la campaña":   None,
        "Comentarios adicionales":  None,
    }])

    numeros_campana = {
        "ID Campaña":               1,
        "Nombre campaña":           2,
        "Número de campaña":        3,
        "Año inicio":               4,
        "Mes inicio":               5,
        "Día inicio":               6,
        "Año término":              7,
        "Mes término":              8,
        "Día término":              9,
        "Objetivo de la campaña":   10,
        "Comentarios adicionales":  11,
    }

    # ───── Hoja "EstacionReplica"
    ESTACION_MAP = {
        "ID Campaña":                  None,
        "Nombre estación":             "name",
        "Tipo de monitoreo":           "tipoMonitoreo",
        "Número Réplica":              None,             # se rellena
        "Descripción EstacionReplica": "comentario",
        "Superficie (m2)":             "tamano",
        "Latitud decimal central":     "coordinatesPlani",
        "Longitud decimal central":    "coordinatesPlani",
        "Región":                      "region",
        "Provincia":                   "provincia",
        "Comuna":                      "comuna",
        "Localidad":                   "localidad",
        "Ecosistema nivel 1":          "cobertura1",
        "Ecosistema nivel 2":          "cobertura2",
        "estacionID":                  "estacionID",
    }
    numeros_estacion = {
        "ID Campaña": 1,
        "Nombre estación": 2,
        "Tipo de monitoreo": 3,
        "Número Réplica": 4,
        "Descripción EstacionReplica": 5,
        "Superficie (m2)": 9,
        "Latitud decimal central": 10,
        "Longitud decimal central": 11,
        "Región": 16,
        "Provincia": 17,
        "Comuna": 18,
        "Localidad": 19,
        "Ecosistema nivel 1": 20,
        "Ecosistema nivel 2": 21,
    }

    df_estacion_columns   = list(ESTACION_MAP.keys())
    df_estacion_plantilla = pd.DataFrame(columns=df_estacion_columns)
    for i, src in df_estacion.reset_index(drop=True).iterrows():
        id_row = i + 1
        for col_name in df_estacion_columns:
            col = ESTACION_MAP.get(col_name)
            if col_name == "ID Campaña":
                cellValue = id_row
            elif col_name == "Número Réplica":
                cellValue = 0   # se rellena con cumcount
            elif col_name == "Latitud decimal central":
                cellValue = get_lat(src[col]) if (col in src and pd.notna(src[col])) else None
            elif col_name == "Longitud decimal central":
                cellValue = get_lon(src[col]) if (col in src and pd.notna(src[col])) else None
            else:
                cellValue = src[col] if (col is not None and col in src and pd.notna(src[col])) else None
            df_estacion_plantilla.loc[i, col_name] = cellValue

    # réplicas por estación
    if not df_estacion_plantilla.empty:
        df_estacion_plantilla["Número Réplica"] = (
            df_estacion_plantilla.groupby("Nombre estación").cumcount().add(1)
        )

    # ───── Hoja "Ocurrencia" (REGISTRO) + join con Número Réplica
    if "estacionID" in df_estacion_plantilla.columns and "estacionID" in df_registro.columns:
        df_estacion_plantilla["estacionID"] = df_estacion_plantilla["estacionID"].astype(str)
        df_registro["estacionID"]           = df_registro["estacionID"].astype(str)
        df_registro = df_registro.merge(
            df_estacion_plantilla[["estacionID", "Número Réplica"]],
            on="estacionID",
            how="left",
        )
    else:
        df_registro["Número Réplica"] = None

    REGISTRO_MAP = {
        "ID Campaña": None,
        "AUTOCOMPLETADO NombreCampaña": "valor",
        "ID EstacionReplica": "Número Réplica",
        "AUTOCOMPLETADO NombreEstacion-Número Replica-Tipo de monitoreo": None,
        "Año del evento": "registroAnoDate",
        "Mes del evento": "registrosMesDate",
        "Día del evento": "registrosDiaDate",
        "Hora inicio evento (hh:mm)": "registrosHoraDate",
        "Protocolo de muestreo": "protocoloMuestreo",
        "Tamaño de la muestra": "tamanoEst",
        "Unidad del tamaño de la muestra": "unidadDeLaMuestra",
        "Esfuerzo de muestreo": None,
        "Profundidad (m)":  None,
        "Comentarios del evento": "comentarios",
        "Reino": "Reino",
        "Filo o división": None,
        "Clase": "clase",
        "Orden":  None,
        "Familia": "familia",
        "Género": "genero",
        "Subgénero": None,
        "Epíteto específico": None,
        "Epíteto infraespecífico": None,
        "Nombre común":  None,
        "Comentarios del taxón":  None,
        "Estado del organismo": "estadoDelOrganismo",
        "Tipo de componente abiótico": "tipoDeComponente",
        "Parámetro": "parametro",
        "Tipo de cuantificación": "tipoCuantificacion",
        "Valor": "nInd",
        "Unidad de valor": "unidadDeValor",
        "Latitud decimal registro": "coordinatesReg",
        "Longitud decimal registro": "coordinatesReg",
        "Hora registro": "registrosHoraDate",
        "Condición reproductiva": None,
        "Sexo (Fauna)": None,
        "Etapa de vida (Fauna)": None,
        "Comportamiento (Fauna)": None,
        "Hábito de crecimiento (Flora)": "habito",
        "Propiedades dinámicas": "valor",
        "Tipo de registro": "tipoDeRegistro",
        "Código individuo": None,
        "Comentarios del registro biológico": None,
        "Muestreado por": None,
        "Identificado por": None,
        "Comentarios de la Identificación": None,
        "Observaciones adicionales": None,
    }

    # Punto E: DEJADO APOSTA con clave 43 duplicada
    numero_registro = {
        1: "ID Campaña",
        2: "AUTOCOMPLETADO NombreCampaña",
        3: "ID EstacionReplica",
        4: "AUTOCOMPLETADO NombreEstacion-Número Replica-Tipo de monitoreo",
        5: "Año del evento",
        6: "Mes del evento",
        7: "Día del evento",
        8: "Hora inicio evento (hh:mm)",
        9: "Protocolo de muestreo",
        10: "Tamaño de la muestra",
        11: "Unidad del tamaño de la muestra",
        12: "Esfuerzo de muestreo",
        13: "Profundidad (m)",
        14: "Comentarios del evento",
        15: "Reino",
        16: "Filo o división",
        17: "Clase",
        18: "Orden",
        19: "Familia",
        20: "Género",
        21: "Subgénero",
        22: "Epíteto específico",
        23: "Epíteto infraespecífico",
        24: "Nombre común",
        25: "Comentarios del taxón",
        26: "Estado del organismo",
        27: "Tipo de componente abiótico",
        28: "Parámetro",
        29: "Tipo de cuantificación",
        30: "Valor",
        31: "Unidad de valor",
        32: "Latitud decimal registro",
        33: "Longitud decimal registro",
        34: "Hora registro",
        35: "Condición reproductiva",
        36: "Sexo (Fauna)",
        37: "Etapa de vida (Fauna)",
        38: "Comportamiento (Fauna)",
        39: "Hábito de crecimiento (Flora)",
        40: "Propiedades dinámicas",
        41: "Tipo de registro",
        42: "Código individuo",
        43: "Comentarios del registro biológico",
        43: "Muestreado por",   # ← duplicado intencional
        45: "Identificado por",
        46: "Comentarios de la Identificación",
        47: "Observaciones adicionales",
    }

    # Construir df_registro_plantilla
    df_registro_plantilla = pd.DataFrame()
    df_registro_columns   = list(REGISTRO_MAP.keys())

    for i in range(len(df_registro)):
        for col_name in df_registro_columns:
            col = REGISTRO_MAP.get(col_name)
            if col_name == "ID Campaña":
                cellValue = 1
            elif col_name == "Latitud decimal registro":
                cellValue = get_lat(df_registro[col][i]) if (col in df_registro and pd.notna(df_registro[col][i])) else None
            elif col_name == "Longitud decimal registro":
                cellValue = get_lon(df_registro[col][i]) if (col in df_registro and pd.notna(df_registro[col][i])) else None
            elif col_name in {"Identificado por", "Comentarios de la Identificación"}:
                cellValue = "AMS Consultores"
            elif col is None:
                # campos sin fuente → dejar vacío (o " " si prefieres)
                cellValue = " " if col_name in {
                    "Esfuerzo de muestreo", "Profundidad (m)", "Filo o división", "Orden", "Subgénero",
                    "Epíteto específico", "Epíteto infraespecífico", "Nombre común", "Comentarios del taxón",
                    "Condición reproductiva", "Sexo (Fauna)", "Etapa de vida (Fauna)", "Comportamiento (Fauna)",
                    "Código individuo", "Comentarios del registro biológico", "Muestreado por",
                    "Identificado por", "Comentarios de la Identificación", "Observaciones adicionales",
                } else None
            else:
                cellValue = df_registro.loc[i, col] if (col in df_registro.columns and pd.notna(df_registro.loc[i, col])) else None

            df_registro_plantilla.loc[i, col_name] = cellValue

    # ───── Escribir a Excel
    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail=f"No se encontró la plantilla: {TEMPLATE_PATH.name}")

    wb = load_workbook(TEMPLATE_PATH)
    wc = wb["Campaña"]
    ws = wb["EstacionReplica"]
    wo = wb["Ocurrencia"]

    # Campaña (fila 3)
    for r in range(3, len(df_campana_plantilla) + 3):
        for col_name, col_number in numeros_campana.items():
            v = df_campana_plantilla.loc[r - 3, col_name]
            wc.cell(row=r, column=col_number, value=v)

    # EstacionReplica (borrar previas y desde fila 2)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for r in range(2, len(df_estacion_plantilla) + 2):
        for col_name, col_number in numeros_estacion.items():
            v = df_estacion_plantilla.loc[r - 2, col_name] if not df_estacion_plantilla.empty else None
            ws.cell(row=r, column=col_number, value=v)

    # Ocurrencia (borrar previas y desde fila 3)
    if wo.max_row > 2:
        wo.delete_rows(3, wo.max_row - 2)
    start_row = 3
    for r in range(start_row, len(df_registro_plantilla) + start_row):
        for col_number in sorted(numero_registro.keys()):
            col_name = numero_registro[col_number]
            v = df_registro_plantilla.loc[r - start_row, col_name] if not df_registro_plantilla.empty else None
            wo.cell(row=r, column=col_number, value=v)

    # Guardar a /tmp/downloads
    out_name = f"Flora_{_safe_filename(campana_id)}_{uuid.uuid4().hex[:6]}.xlsx"
    out_path = DOWNLOAD_DIR / out_name
    wb.save(out_path)
    return out_path

# ──────────────────────────────── Endpoints
@app.get("/health")
def health():
    return {"ok": True}

# ✅ Endpoint de descarga con Content-Disposition (mejor para iOS)
@app.get("/download/{fname}")
def download_file(fname: str):
    file_path = DOWNLOAD_DIR / fname
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no existe")
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,  # fuerza Content-Disposition: attachment
        headers={
            "Access-Control-Expose-Headers": "Content-Disposition",
            "Cache-Control": "no-store",
        },
    )

@app.get("/export")
def export_excel(request: Request, campana_id: str = Query(..., description="campanaID a exportar")):
    try:
        out_path = generar_excel_fauna_like(campana_id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {e}")

    # ✅ URL absoluta correcta detrás de proxy y en https
    download_url = request.url_for("download_file", fname=out_path.name)
    return JSONResponse({"download_url": str(download_url)})


